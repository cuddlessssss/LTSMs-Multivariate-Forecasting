import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.preprocessing import LabelEncoder
import tensorflow as tf
from tensorflow.keras.models import Model
from tensorflow.keras.layers import LSTM, Dense, Input
from tensorflow.keras.optimizers import Adam
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# 1. 🗂️ Load Data
data = pd.read_excel('your_past_data.xlsx', sheet_name='updated')
data.columns = data.columns.str.strip()

# Remove rows with negative sell-out qty
data = data[data['Dealer Net Sell Out Qty'] >= 0]

data['Date'] = pd.to_datetime(data['Date'], errors='coerce')
data['year_month'] = data['Date'].dt.to_period('M')
latest_date = data['Date'].max()
cutoff_date = latest_date - pd.DateOffset(years=4)
data = data[data['Date'] >= cutoff_date]

# 🧾 Load future total sell-outs in pivoted format
future_total_raw = pd.read_excel('future_total_sell_outs_pivoted.xlsx')
future_total_melted = future_total_raw.melt(id_vars='Model Name', var_name='year_month', value_name='total_sell_out')
future_total_melted['year_month'] = pd.to_datetime(future_total_melted['year_month']).dt.to_period('M')

# 2. ⚙️ Preprocessing
f1_encoder = LabelEncoder()
f2_encoder = LabelEncoder()
data['Model Name_enc'] = f1_encoder.fit_transform(data['Model Name'])
data['Account Name_enc'] = f2_encoder.fit_transform(data['Account Name'])

grouped = data.groupby(['year_month', 'Model Name_enc', 'Account Name_enc']).agg({'Dealer Net Sell Out Qty': 'sum'}).reset_index()
pivot = grouped.pivot_table(index='year_month', columns=['Model Name_enc', 'Account Name_enc'], values='Dealer Net Sell Out Qty', fill_value=0)
pivot = pivot.sort_index(axis=1)
pivot = pivot.div(pivot.sum(axis=1), axis=0)

# Prepare sequences
input_sequence_length = 12
X, y = [], []
if len(pivot) > input_sequence_length:
    for i in range(input_sequence_length, len(pivot)):
        X.append(pivot.iloc[i - input_sequence_length:i].values)
        y.append(pivot.iloc[i].values)
    X = np.array(X)
    y = np.array(y)
else:
    raise ValueError("Not enough data for training")

# 3. 🧠 Build Model
input_layer = Input(shape=(input_sequence_length, X.shape[2]))
lstm_out = LSTM(128)(input_layer)
output = Dense(X.shape[2], activation='softmax')(lstm_out)
model = Model(inputs=input_layer, outputs=output)
model.compile(loss='categorical_crossentropy', optimizer=Adam(learning_rate=0.001))
model.fit(X, y, epochs=100, batch_size=16)

# 4. 📈 Forecasting
last_sequence = pivot.iloc[-input_sequence_length:].values
future_preds = []
column_index = pivot.columns

for i in range(12):
    pred = model.predict(last_sequence[np.newaxis, :, :])[0]
    future_preds.append(pred)
    last_sequence = np.vstack([last_sequence[1:], pred])

# 5. 🔄 Rescale + Round
rounded_predictions = []
highlight_map = []

unique_forecast_months = future_total_melted['year_month'].unique()

for i, pred in enumerate(future_preds):
    forecast_month = unique_forecast_months[i]

    # Step 1: Create a dataframe for the predicted proportions
    pred_df = pd.DataFrame({
        'model_enc': [col[0] for col in column_index],
        'account_enc': [col[1] for col in column_index],
        'proportion': pred
    })
    pred_df['model_name'] = f1_encoder.inverse_transform(pred_df['model_enc'])

    # Step 2: Get the model-level forecast totals
    model_total_map = future_total_melted[future_total_melted['year_month'] == forecast_month] \
        .set_index('Model Name')['total_sell_out'].to_dict()

    # Step 3: Filter only those models that have forecast totals
    pred_df = pred_df[pred_df['model_name'].isin(model_total_map.keys())]

    # Step 4: Normalize proportions per model
    pred_df['total_model_qty'] = pred_df['model_name'].map(model_total_map)
    pred_df['model_sum'] = pred_df.groupby('model_name')['proportion'].transform('sum')
    pred_df['scaled_qty'] = pred_df['proportion'] / pred_df['model_sum'] * pred_df['total_model_qty']

    # Step 5: Convert back to raw_pred array (aligned with column_index)
    raw_pred = np.zeros(len(column_index))
    highlight_status = [""] * len(column_index)

    decimals = []
    for idx, (model_enc, account_enc) in enumerate(column_index):
        match = pred_df[(pred_df['model_enc'] == model_enc) & (pred_df['account_enc'] == account_enc)]
        if not match.empty:
            qty = match['scaled_qty'].values[0]
            raw_pred[idx] = qty
            decimals.append(qty - np.floor(qty))
        else:
            decimals.append(0)

    # Step 6: Round while tracking status
    df_temp = pd.DataFrame({
        'value': raw_pred,
        'decimals': decimals,
        'col_idx': range(len(raw_pred))
    }).sort_values('decimals')

    n = len(df_temp)
    down_set = df_temp.iloc[:n // 2]
    up_set = df_temp.iloc[n // 2:]

    rounded = np.zeros_like(raw_pred)
    round_status = [""] * len(raw_pred)
    for _, row in down_set.iterrows():
        idx = int(row['col_idx'])
        rounded[idx] = np.floor(raw_pred[idx])
        round_status[idx] = "down"
    for _, row in up_set.iterrows():
        idx = int(row['col_idx'])
        rounded[idx] = np.ceil(raw_pred[idx])
        round_status[idx] = "up"

    # Step 7: Global adjustment if rounding error exists
    diff = int(round(np.sum(raw_pred) - np.sum(rounded)))
    if diff != 0:
        adjust_idxs = np.argsort(decimals)
        if diff > 0:
            for j in adjust_idxs[::-1]:
                rounded[j] += 1
                round_status[j] = "up"
                diff -= 1
                if diff == 0: break
        else:
            for j in adjust_idxs:
                if rounded[j] > 0:
                    rounded[j] -= 1
                    round_status[j] = "down"
                    diff += 1
                    if diff == 0: break

    # Save rounded result and highlight map
    rounded_predictions.append(rounded)
    highlight_map.append(round_status)

    decimals = raw_pred - np.floor(raw_pred)
    df_temp = pd.DataFrame({
        'value': raw_pred,
        'decimals': decimals,
        'col_idx': range(len(raw_pred))
    }).sort_values('decimals')

    n = len(df_temp)
    down_set = df_temp.iloc[:n // 2]
    up_set = df_temp.iloc[n // 2:]

    rounded = np.zeros_like(raw_pred)
    round_status = [""] * len(raw_pred)
    for _, row in down_set.iterrows():
        idx = int(row['col_idx'])
        rounded[idx] = np.floor(raw_pred[idx])
        round_status[idx] = "down"
    for _, row in up_set.iterrows():
        idx = int(row['col_idx'])
        rounded[idx] = np.ceil(raw_pred[idx])
        round_status[idx] = "up"

    diff = int(round(np.sum(raw_pred) - np.sum(rounded)))
    if diff != 0:
        adjust_idxs = np.argsort(decimals)
        if diff > 0:
            for j in adjust_idxs[::-1]:
                rounded[j] += 1
                round_status[j] = "up"
                diff -= 1
                if diff == 0: break
        else:
            for j in adjust_idxs:
                if rounded[j] > 0:
                    rounded[j] -= 1
                    round_status[j] = "down"
                    diff += 1
                    if diff == 0: break

    rounded_predictions.append(rounded)
    highlight_map.append(round_status)

rounded_predictions = np.array(rounded_predictions)

# 6. 📊 Prepare result dataframe
result_records = []
for month_idx in range(12):
    forecast_month = str(unique_forecast_months[month_idx])
    for col_idx, value in enumerate(rounded_predictions[month_idx]):
        feature_1_enc, feature_2_enc = column_index[col_idx]
        model_name = f1_encoder.inverse_transform([feature_1_enc])[0]
        account_name = f2_encoder.inverse_transform([feature_2_enc])[0]
        result_records.append({
            'Model Name': model_name,
            'Account Name': account_name,
            'Month': forecast_month,
            'Qty': value,
            'Highlight': highlight_map[month_idx][col_idx]
        })

result_df = pd.DataFrame(result_records)

# Pivot to wide format
pivoted = result_df.pivot_table(index=['Model Name', 'Account Name'], columns='Month', values='Qty', fill_value=0).reset_index()

# 7. 📥 Save to Excel with highlights and legend
wb = Workbook()
ws = wb.active
ws.title = "Predictions_Pivoted"

for r in dataframe_to_rows(pivoted, index=False, header=True):
    ws.append(r)

red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

header_row = 1
month_columns = pivoted.columns[2:]

for row_idx in range(2, ws.max_row + 1):
    model_name = ws.cell(row=row_idx, column=1).value
    account_name = ws.cell(row=row_idx, column=2).value
    for col_offset, month in enumerate(month_columns):
        col_idx = col_offset + 3
        highlight_row = result_df[
            (result_df['Model Name'] == model_name) &
            (result_df['Account Name'] == account_name) &
            (result_df['Month'] == month)
        ]
        if not highlight_row.empty:
            status = highlight_row['Highlight'].values[0]
            if status == 'up':
                ws.cell(row=row_idx, column=col_idx).fill = red_fill
            elif status == 'down':
                ws.cell(row=row_idx, column=col_idx).fill = green_fill

# Add legend
ws.insert_rows(1)
ws.cell(row=1, column=1).value = "Legend:"
ws.cell(row=1, column=2).value = "Rounded Up"
ws.cell(row=1, column=3).value = "Rounded Down"
ws.cell(row=1, column=2).fill = red_fill
ws.cell(row=1, column=3).fill = green_fill

wb.save("Predictions_Pivoted2.xlsx")
print("✅ Done! Excel with pivoted predictions, highlights, and per-model scaling saved.")